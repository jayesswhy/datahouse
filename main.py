from openpyxl.styles import PatternFill
from openpyxl import load_workbook
wb = load_workbook('names.xlsx')
sheet = wb['Sheet1']

# iterate through worksheet and compare first cell with other cells in row to see if it matches
for row in sheet.iter_rows():
    true_value = row[0]
    for cell in row:
       if cell.value != true_value.value:
          cell.fill = PatternFill(start_color="feedc6", end_color="feedc6", fill_type="solid")
# creates a spreadsheet with highlighted cells and non-highlighted cells
wb.save('names_revised1.xlsx')

